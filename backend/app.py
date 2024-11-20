from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
import os
import bcrypt
import logging
from marshmallow import Schema, fields, ValidationError
import openpyxl
from datetime import datetime

app = Flask(__name__)

# Configurar el logging
logging.basicConfig(level=logging.DEBUG)

# Configurar CORS para permitir solicitudes desde http://localhost:50735
CORS(app, resources={
    r"/register": {"origins": "http://localhost:50735", "allow_headers": ["Content-Type"], "supports_credentials": True},
    r"/login": {"origins": "http://localhost:50735", "allow_headers": ["Content-Type"], "supports_credentials": True},
    r"/sales": {"origins": "http://localhost:50735", "allow_headers": ["Content-Type"], "supports_credentials": True},
    r"/search_menu": {"origins": "http://localhost:50735", "allow_headers": ["Content-Type"], "supports_credentials": True},
    r"/add_item": {"origins": "http://localhost:50735", "allow_headers": ["Content-Type"], "supports_credentials": True}
})

# Conectar a la base de datos
def get_db_connection():
    conn = psycopg2.connect(
        dbname=os.getenv('DATABASE_NAME'),
        user=os.getenv('DATABASE_USER'),
        password=os.getenv('DATABASE_PASSWORD'),
        host=os.getenv('DATABASE_HOST'),
        port=os.getenv('DATABASE_PORT')
    )
    return conn

# Validar las credenciales de administrador
def validate_admin(admin_username, admin_password):
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    admin = cursor.execute('SELECT * FROM users WHERE username = %s', (admin_username,)).fetchone()
    conn.close()
    if admin and bcrypt.checkpw(admin_password.encode('utf-8'), admin['password']):
        return True
    return False

# Esquema de validación para el registro de usuarios
class UserSchema(Schema):
    username = fields.Str(required=True)
    password = fields.Str(required=True)
    admin_username = fields.Str()
    admin_password = fields.Str()

def validate_data(schema, data):
    try:
        schema.load(data)
    except ValidationError as err:
        logging.error(f"Validation error: {err.messages}")
        return False
    return True

# Ruta para registrar un nuevo usuario
@app.route('/register', methods=['POST', 'OPTIONS'])
def register_user():
    if request.method == 'OPTIONS':
        logging.debug("Handling OPTIONS request for /register")
        response = jsonify({'message': 'Preflight request handled'})
        return add_cors_headers(response)

    data = request.json
    if not validate_data(UserSchema(), data):
        response = jsonify({'message': 'Datos inválidos'}), 400
        return add_cors_headers(response)

    username = data['username']
    password = data['password']
    admin_username = data.get('admin_username')
    admin_password = data.get('admin_password')

    logging.debug(f"Received registration request for user: {username}")

    if not validate_admin(admin_username, admin_password):
        logging.warning("Invalid admin credentials provided")
        response = jsonify({'message': 'Credenciales de administrador inválidas'}), 403
        return add_cors_headers(response)

    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO users (username, password) VALUES (%s, %s)', (username, hashed_password))
        conn.commit()
        response = jsonify({'message': 'Usuario registrado'}), 201
    except psycopg2.IntegrityError:
        logging.warning(f"User {username} already exists")
        response = jsonify({'message': 'Usuario ya existe'}), 400
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        response = jsonify({'message': 'Error interno del servidor'}), 500
    finally:
        cursor.close()
        conn.close()

    return add_cors_headers(response)

# Ruta para iniciar sesión
@app.route('/login', methods=['POST', 'OPTIONS'])
def login_user():
    if request.method == 'OPTIONS':
        logging.debug("Handling OPTIONS request for /login")
        response = jsonify({'message': 'Preflight request handled'})
        return add_cors_headers(response)

    data = request.json
    username = data['username']
    password = data['password']

    logging.debug(f"Received login request for user: {username}")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    user = cursor.execute('SELECT * FROM users WHERE username = %s', (username,)).fetchone()
    conn.close()

    if user and bcrypt.checkpw(password.encode('utf-8'), user['password']):
        logging.info(f"User {username} logged in successfully")
        response = jsonify({'message': 'Inicio de sesión exitoso', 'username': username}), 200
    else:
        logging.warning(f"Invalid credentials for user: {username}")
        response = jsonify({'message': 'Credenciales inválidas'}), 401

    return add_cors_headers(response)

# Ruta para registrar una venta
@app.route('/sales', methods=['POST', 'OPTIONS'])
def add_sale():
    if request.method == 'OPTIONS':
        logging.debug("Handling OPTIONS request for /sales")
        response = jsonify({'message': 'Preflight request handled'})
        return add_cors_headers(response)

    data = request.json
    if 'items' not in data or 'username' not in data or 'timestamp' not in data:
        logging.error("Missing 'items', 'username', or 'timestamp' in request body")
        response = jsonify({'message': 'Falta el campo "items", "username" o "timestamp" en la solicitud'}), 400
        return add_cors_headers(response)

    items = data['items']
    username = data['username']
    timestamp = data['timestamp']

    logging.debug(f"Received add sale request for items: {items} by user: {username} at {timestamp}")

    conn = get_db_connection()
    cursor = conn.cursor()
    total_price = 0.0
    try:
        for item in items:
            total_item_price = item['quantity'] * item['price']
            total_price += total_item_price
            cursor.execute('INSERT INTO sales (username, item, quantity, price, total_price, timestamp) VALUES (%s, %s, %s, %s, %s, %s)',
                           (username, item['name'], item['quantity'], item['price'], total_item_price, timestamp))
        conn.commit()
        response = jsonify({'message': 'Venta registrada', 'total_price': total_price}), 201
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        response = jsonify({'message': 'Error interno del servidor'}), 500
    finally:
        cursor.close()
        conn.close()

    # Guardar en Excel
    save_to_excel(items, username, timestamp)

    return add_cors_headers(response)

# Ruta para obtener las ventas
@app.route('/sales', methods=['GET', 'OPTIONS'])
def get_sales():
    if request.method == 'OPTIONS':
        logging.debug("Handling OPTIONS request for /sales")
        response = jsonify({'message': 'Preflight request handled'})
        return add_cors_headers(response)

    logging.debug("Fetching sales data")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        sales = cursor.execute('SELECT * FROM sales ORDER BY timestamp DESC').fetchall()
        daily_sales = []
        daily_total = 0.0
        current_date = None

        for sale in sales:
            sale_date = sale['timestamp'].split(' ')[0]
            if current_date != sale_date:
                if current_date is not None:
                    daily_sales.append({
                        'date': current_date,
                        'total': daily_total
                    })
                current_date = sale_date
                daily_total = 0.0

            daily_sales.append({
                'id': sale['id'],
                'username': sale['username'],
                'item': sale['item'],
                'quantity': sale['quantity'],
                'price': sale['price'],
                'total_price': sale['total_price'],
                'timestamp': sale['timestamp']
            })
            daily_total += sale['total_price']

        if current_date is not None:
            daily_sales.append({
                'date': current_date,
                'total': daily_total
            })

        response = jsonify(daily_sales), 200
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        response = jsonify({'message': 'Error interno del servidor'}), 500
    finally:
        cursor.close()
        conn.close()

    return add_cors_headers(response)

# Ruta para buscar platos en el menú
@app.route('/search_menu', methods=['GET', 'OPTIONS'])
def search_menu():
    if request.method == 'OPTIONS':
        logging.debug("Handling OPTIONS request for /search_menu")
        response = jsonify({'message': 'Preflight request handled'})
        return add_cors_headers(response)

    query = request.args.get('query', '')

    logging.debug(f"Received search request for query: {query}")

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        results = cursor.execute('SELECT * FROM menu WHERE name LIKE %s', (f'%{query}%',)).fetchall()
        response = jsonify([dict(row) for row in results]), 200
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        response = jsonify({'message': 'Error interno del servidor'}), 500
    finally:
        cursor.close()
        conn.close()

    return add_cors_headers(response)

# Ruta para agregar un nuevo item al menú
@app.route('/add_item', methods=['POST', 'OPTIONS'])
def add_item():
    if request.method == 'OPTIONS':
        logging.debug("Handling OPTIONS request for /add_item")
        response = jsonify({'message': 'Preflight request handled'})
        return add_cors_headers(response)

    data = request.json
    if 'name' not in data or 'price' not in data:
        logging.error("Missing 'name' or 'price' in request body")
        response = jsonify({'message': 'Falta el campo "name" o "price" en la solicitud'}), 400
        return add_cors_headers(response)

    name = data['name']
    price = data['price']

    logging.debug(f"Received add item request for item: {name} with price: {price}")

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO menu (name, price) VALUES (%s, %s)', (name, price))
        conn.commit()
        response = jsonify({'message': 'Item added successfully'}), 201
    except psycopg2.IntegrityError:
        logging.warning(f"Item {name} already exists")
        response = jsonify({'message': 'Item already exists'}), 400
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        response = jsonify({'message': 'Error interno del servidor'}), 500
    finally:
        cursor.close()
        conn.close()

    return add_cors_headers(response)

# Ruta para ver las ventas en una página HTML
@app.route('/view_sales', methods=['GET'])
def view_sales():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    try:
        sales = cursor.execute('SELECT * FROM sales ORDER BY timestamp DESC').fetchall()
        return render_template('sales.html', sales=sales)
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        return "Error interno del servidor", 500
    finally:
        cursor.close()
        conn.close()

# Función para guardar las ventas en un archivo Excel
def save_to_excel(items, username, timestamp):
    date = datetime.fromisoformat(timestamp)
    month_str = date.strftime('%B %Y')
    day_str = date.strftime('%d %B %Y')
    user_documents_path = os.path.expanduser('~\\Documents\\')
    sales_folder_path = os.path.join(user_documents_path, 'Restaurant Sales')
    
    # Crear la carpeta si no existe
    if not os.path.exists(sales_folder_path):
        os.makedirs(sales_folder_path)

    file_path = os.path.join(sales_folder_path, f'sales_{month_str}.xlsx')

    # Crear un nuevo libro de Excel si no existe
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Crear una hoja nueva para el día si no existe
    if day_str not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=day_str)
        sheet.append(['ID', 'Usuario', 'Item', 'Cantidad', 'Precio', 'Total', 'Timestamp'])
    else:
        sheet = workbook[day_str]

    # Agregar las ventas a la hoja
    for item in items:
        total_item_price = item['quantity'] * item['price']
        sheet.append([
            None,  # ID será autoincremental en la base de datos
            username,
            item['name'],
            item['quantity'],
            item['price'],
            total_item_price,
            timestamp
        ])

    # Guardar el archivo Excel
    workbook.save(file_path)
    logging.debug(f"Sales saved to Excel file: {file_path}")

# Función para manejar el encabezado Access-Control-Allow-Origin dinámicamente
def add_cors_headers(response):
    if isinstance(response, tuple):
        response, status_code = response
    else:
        status_code = 200

    if 'Origin' in request.headers:
        response.headers['Access-Control-Allow-Origin'] = request.headers['Origin']
    response.headers['Access-Control-Allow-Methods'] = 'POST, GET, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    logging.debug(f"Response headers: {response.headers}")
    return response, status_code

# Manejo de errores
@app.errorhandler(400)
def bad_request(error):
    logging.error(f"Bad request: {error}")
    response = jsonify({'message': 'Solicitud incorrecta'}), 400
    return add_cors_headers(response)

@app.errorhandler(401)
def unauthorized(error):
    logging.error(f"Unauthorized: {error}")
    response = jsonify({'message': 'No autorizado'}), 401
    return add_cors_headers(response)

@app.errorhandler(403)
def forbidden(error):
    logging.error(f"Forbidden: {error}")
    response = jsonify({'message': 'Prohibido'}), 403
    return add_cors_headers(response)

@app.errorhandler(500)
def internal_server_error(error):
    logging.error(f"Internal server error: {error}")
    response = jsonify({'message': 'Error interno del servidor'}), 500
    return add_cors_headers(response)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)